"""
Excel template generation + parsing/validation for roster and result
uploads. Kept separate from views.py so the parsing logic can be unit
tested without touching HTTP at all.
"""

from collections import Counter
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from django.core.exceptions import ValidationError
from openpyxl import Workbook

from apps.accounts.models import reg_number_validator

#MIN_ASSESSMENT_COLUMNS = 6
MIN_FILLED_SCORES = 3
ROSTER_HEADERS = ["Reg No", "Name"]
DEFAULT_CA_TOTAL = 40


def _is_blank(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def _normalize_reg_number(value):
    return str(value).strip().upper()


# ── Template generation ──────────────────────────────────────────────────

def generate_roster_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(ROSTER_HEADERS)
    ws.append(["PS/CSC/22/0001", "Ama Owusu"])
    return wb


def generate_result_template():
    """
    Generic example template. Row 1 = column names, Row 2 = max score for
    each assessment column (blank under Reg No / Name), Row 3+ = data.

    Two of the six example columns are left with a BLANK max score on
    purpose, to show that not every column needs to be graded in a given
    upload — a lecturer can grade 2 (or 3, or any number) of the columns
    now and leave the rest for a later upload. A blank-max column is
    excluded from the CA calculation entirely, for every student, until a
    max score is filled in for it.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    columns = [
        ("A1", 5), ("Quiz", 10), ("Attendance", None),
        ("Midsem", 10), ("Project", None), ("Test", 5),
    ]
    ws.append(["Reg No", "Name"] + [c[0] for c in columns])
    ws.append(["", ""] + [c[1] if c[1] is not None else "" for c in columns])
    ws.append(["PS/CSC/22/0001", "Ama Owusu"] + [
        0 if max_score is not None else "" for _, max_score in columns
    ])
    return wb


def workbook_to_bytes(wb):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Roster parsing ───────────────────────────────────────────────────────

def parse_roster_file(file_obj):
    """
    Returns (valid_rows, errors, file_rejected).
    valid_rows: list of {"reg_number", "name"}
    file_rejected: True if the whole file was thrown out (bad headers or
    a duplicate reg number anywhere in the file).
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception:
        return [], ["Could not read this file. Make sure it's a valid .xlsx file."], True

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    rows = [r for r in rows if r is not None and not all(_is_blank(v) for v in r)]

    if not rows:
        return [], ["The file is empty."], True

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if headers[:2] != ROSTER_HEADERS:
        return [], [
            f"The first two columns must be exactly 'Reg No' and 'Name'. "
            f"Found: {headers[:2]}"
        ], True

    data_rows = rows[1:]

    raw_reg_numbers = [
        _normalize_reg_number(r[0]) for r in data_rows if not _is_blank(r[0])
    ]
    dupes = [rn for rn, count in Counter(raw_reg_numbers).items() if count > 1]
    if dupes:
        return [], [
            f"Duplicate registration number(s) found in the file: "
            f"{', '.join(dupes)}. Whole file rejected — please fix and re-upload."
        ], True

    valid_rows = []
    errors = []

    for i, row in enumerate(data_rows, start=2):
        reg_number = row[0] if len(row) > 0 else None
        name = row[1] if len(row) > 1 else None

        if _is_blank(reg_number):
            errors.append(f"Row {i}: missing registration number — row skipped.")
            continue
        reg_number = _normalize_reg_number(reg_number)

        try:
            reg_number_validator(reg_number)
        except ValidationError:
            errors.append(
                f"Row {i}: '{reg_number}' is not a valid format "
                f"(expected PS/CSC/22/0001) — row skipped."
            )
            continue

        if _is_blank(name):
            errors.append(f"Row {i}: missing name for {reg_number} — row skipped.")
            continue

        valid_rows.append({"reg_number": reg_number, "name": str(name).strip()})

    return valid_rows, errors, False


# ── Result parsing ───────────────────────────────────────────────────────

def parse_result_file(file_obj, known_reg_numbers=None, ca_total=DEFAULT_CA_TOTAL):
    """
    known_reg_numbers: set of reg numbers already on the course roster.

    ca_total: the target CA total every student's score is scaled to
    (e.g. 40). Scaling formula: (raw_total / declared_max_total) * ca_total.

    declared_max_total is the sum of max scores for columns that HAVE a
    max score in row 2 — the SAME denominator for every student in the
    upload. A column left blank in row 2 means "not graded in this
    upload": it's excluded entirely, for every student, no matter what's
    typed under it. This lets a lecturer grade only 2 of 6 columns now
    and the rest later, with everyone still landing on a fair /ca_total
    today.

    (A student individually missing one of the ACTIVE columns just gets
    a lower raw_total — the denominator doesn't shrink for them alone,
    since it's fixed at the column level for the whole upload.)

    Returns (assessment_columns, valid_rows, errors, file_rejected).
    assessment_columns: list of (name, Decimal max_score OR None if unused)
    valid_rows: list of dicts with reg_number, name, scores,
                raw_total, raw_max, ca_score, ca_total, low_fill_warning
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception:
        return [], [], ["Could not read this file. Make sure it's a valid .xlsx file."], True

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    rows = [r for r in rows if r is not None and not all(_is_blank(v) for v in r)]

    if len(rows) < 3:
        return [], [], [
            "The file must have a header row, a max-score row, and at "
            "least one data row."
        ], True

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if headers[:2] != ["Reg No", "Name"]:
        return [], [], [
            f"The first two columns must be exactly 'Reg No' and 'Name'. "
            f"Found: {headers[:2]}"
        ], True

    assessment_names = [h for h in headers[2:] if h]
    if not assessment_names:
        return [], [], [
            "At least one assessment column is required (e.g. 'Quiz', "
            "'Midsem')."
        ], True

    # Scores are stored keyed by column NAME, so two columns sharing a
    # name would silently collide — the second one's value would
    # overwrite the first for every student, with no error and no sign
    # anything was lost. Reject outright instead of corrupting data.
    name_counts = Counter(assessment_names)
    duplicate_names = [name for name, count in name_counts.items() if count > 1]
    if duplicate_names:
        return [], [], [
            f"Duplicate column name(s) found: {', '.join(duplicate_names)}. "
            f"Each assessment column must have a unique name — e.g. use "
            f"'Quiz 1' and 'Quiz 2' instead of two columns both named "
            f"'Quiz'. Whole file rejected."
        ], True

    # Row 2 = max score per column. A BLANK cell here means "not graded
    # in this upload" — the whole column is excluded from scoring, no
    # matter what's typed under it. A non-blank cell must still be a
    # positive number, or the file is rejected (that's a data error, not
    # an intentional skip).
    max_score_row = rows[1]
    assessment_columns = []
    for idx, name in enumerate(assessment_names):
        raw_max = max_score_row[2 + idx] if len(max_score_row) > 2 + idx else None
        if _is_blank(raw_max):
            assessment_columns.append((name, None))
            continue
        try:
            max_score = Decimal(str(raw_max))
            if max_score <= 0:
                raise InvalidOperation
        except (InvalidOperation, TypeError, ValueError):
            return [], [], [
                f"Column '{name}': row 2 (max score) must be either blank "
                f"(column not graded yet) or a positive number — found "
                f"'{raw_max}'. Whole file rejected."
            ], True
        assessment_columns.append((name, max_score))

    active_columns = [(n, m) for n, m in assessment_columns if m is not None]
    if not active_columns:
        return [], [], [
            "At least one assessment column needs a max score in row 2 "
            "— every column is currently blank."
        ], True

    declared_max_total = sum(m for _, m in active_columns)
    num_active_columns = len(active_columns)

    data_rows = rows[2:]

    # File-wide duplicate check, same reasoning as the roster parser.
    raw_reg_numbers = [
        _normalize_reg_number(r[0]) for r in data_rows if not _is_blank(r[0])
    ]
    dupes = [rn for rn, count in Counter(raw_reg_numbers).items() if count > 1]
    if dupes:
        return [], [], [
            f"Duplicate registration number(s) found in the file: "
            f"{', '.join(dupes)}. Whole file rejected — please fix and re-upload."
        ], True

    valid_rows = []
    errors = []

    for i, row in enumerate(data_rows, start=3):
        reg_number = row[0] if len(row) > 0 else None
        name = row[1] if len(row) > 1 else None

        if _is_blank(reg_number):
            errors.append(f"Row {i}: missing registration number — row skipped.")
            continue
        reg_number = _normalize_reg_number(reg_number)

        try:
            reg_number_validator(reg_number)
        except ValidationError:
            errors.append(f"Row {i}: '{reg_number}' is not a valid format — row skipped.")
            continue

        if _is_blank(name):
            errors.append(f"Row {i}: missing name for {reg_number} — row skipped.")
            continue
        name = str(name).strip()

        if known_reg_numbers is not None and reg_number not in known_reg_numbers:
            errors.append(
                f"Row {i}: {reg_number} is not on this course's roster — row skipped."
            )
            continue

        scores = {}
        row_valid = True
        filled_count = 0
        raw_total = Decimal("0")

        for idx, (col_name, max_score) in enumerate(assessment_columns):
            if max_score is None:
                # Column not in use this upload — always blank for every
                # student, regardless of anything typed under it.
                scores[col_name] = None
                continue

            raw_value = row[2 + idx] if len(row) > 2 + idx else None
            if _is_blank(raw_value):
                scores[col_name] = None
                continue
            try:
                score = Decimal(str(raw_value))
            except (InvalidOperation, TypeError, ValueError):
                errors.append(
                    f"Row {i} ({reg_number}): '{raw_value}' in column "
                    f"'{col_name}' is not numeric — row skipped."
                )
                row_valid = False
                break
            if score < 0 or score > max_score:
                errors.append(
                    f"Row {i} ({reg_number}): score {score} in '{col_name}' "
                    f"is outside 0–{max_score} — row skipped."
                )
                row_valid = False
                break
            scores[col_name] = float(score)
            filled_count += 1
            raw_total += score

        if not row_valid:
            continue

        ca_score = (
            round(float(raw_total) / float(declared_max_total) * float(ca_total), 1)
            if declared_max_total
            else 0.0
        )

        # A student needs at least 3 filled columns to avoid the warning —
        # but only 3 out of however many columns are actually active this
        # upload. If only 2 columns are graded at all, filling both is a
        # complete submission and shouldn't be flagged as incomplete.
        low_fill_threshold = min(MIN_FILLED_SCORES, num_active_columns)
        low_fill_warning = filled_count < low_fill_threshold

        valid_rows.append({
            "reg_number": reg_number,
            "name": name,
            "scores": scores,
            "raw_total": float(raw_total),
            "raw_max": float(declared_max_total),
            "ca_score": ca_score,
            "ca_total": float(ca_total),
            "low_fill_warning": low_fill_warning,
        })

    return assessment_columns, valid_rows, errors, False